import os
import re
import json
import math
import asyncio
import secrets
from datetime import datetime, timedelta, timezone, date
from typing import Any, Dict, List, Optional, Set, Tuple
from io import BytesIO

import httpx
import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

# Discord signature verification
from nacl.signing import VerifyKey
from nacl.exceptions import BadSignatureError

# ----------------------------
# Timezone
# ----------------------------
KST = timezone(timedelta(hours=9))

# ----------------------------
# Environment variables
# ----------------------------
DISCORD_CLIENT_ID = os.getenv("DISCORD_CLIENT_ID", "")
DISCORD_CLIENT_SECRET = os.getenv("DISCORD_CLIENT_SECRET", "")
DISCORD_BOT_TOKEN = os.getenv("DISCORD_BOT_TOKEN", "")
DISCORD_PUBLIC_KEY = os.getenv("DISCORD_PUBLIC_KEY", "")

# Guild is required for:
# - Showing server display name reliably
# - (deprecated) Previously used for per-place visit roles
DISCORD_GUILD_ID = int(os.getenv("DISCORD_GUILD_ID", "0") or 0)

# Per-place admin channel can override this global fallback
DISCORD_ADMIN_CHANNEL_ID = os.getenv("DISCORD_ADMIN_CHANNEL_ID", "")  # (fallback) channel id

# Role-gated permissions (CSV of role ids). Backward compatible with single-id vars.
DISCORD_QR_CREATE_ROLE_IDS = os.getenv("DISCORD_QR_CREATE_ROLE_IDS", "") or os.getenv("DISCORD_QR_CREATE_ROLE_ID", "")
DISCORD_ADMIN_ROLE_IDS = os.getenv("DISCORD_ADMIN_ROLE_IDS", "") or os.getenv("DISCORD_ADMIN_ROLE_ID", "")

# Visit log export permissions (CSV of role ids). Backward compatible with single-id var.
DISCORD_EXPORT_ROLE_IDS = os.getenv("DISCORD_EXPORT_ROLE_IDS", "") or os.getenv("DISCORD_EXPORT_ROLE_ID", "")

def _required_env(name: str) -> str:
    value = (os.getenv(name, "") or "").strip()
    if not value:
        raise RuntimeError(f"{name} is required")
    return value

SESSION_SECRET = _required_env("SESSION_SECRET")
HTTPS_ONLY = os.getenv("HTTPS_ONLY", "true").lower() == "true"
OAUTH_REDIRECT_URI = os.getenv("OAUTH_REDIRECT_URI", "")
DATABASE_URL = os.getenv("DATABASE_URL", "")
PUBLIC_BASE_URL = _required_env("PUBLIC_BASE_URL").rstrip("/")

if not re.fullmatch(r"https?://[^\s/$.?#].[^\s]*", PUBLIC_BASE_URL):
    raise RuntimeError("PUBLIC_BASE_URL must be an absolute http(s) URL")
if HTTPS_ONLY and not PUBLIC_BASE_URL.startswith("https://"):
    raise RuntimeError("PUBLIC_BASE_URL must use https:// when HTTPS_ONLY=true")

WEB_SESSION_TTL_SECONDS = int(os.getenv("WEB_SESSION_TTL_SECONDS", "180") or 180)
OAUTH_STATE_TTL_SECONDS = int(os.getenv("OAUTH_STATE_TTL_SECONDS", "600") or 600)
DISCORD_SIGNATURE_MAX_AGE_SECONDS = int(os.getenv("DISCORD_SIGNATURE_MAX_AGE_SECONDS", "300") or 300)

# ----------------------------
# App setup
# ----------------------------
app = FastAPI()
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    https_only=HTTPS_ONLY,
    same_site="lax",
)

templates = Jinja2Templates(directory="templates")

_db_pool: Optional[asyncpg.Pool] = None

# Guild roles cache (id->name)
_roles_cache: Dict[str, Any] = {"ts": 0.0, "map": {}}

def _parse_id_set(raw: str) -> Set[int]:
    out: Set[int] = set()
    for part in re.split(r"[,\s]+", raw or ""):
        part = (part or "").strip()
        if not part:
            continue
        try:
            out.add(int(part))
        except ValueError:
            continue
    return out

QR_CREATE_ROLES = _parse_id_set(DISCORD_QR_CREATE_ROLE_IDS)
ADMIN_ROLES = _parse_id_set(DISCORD_ADMIN_ROLE_IDS)

EXPORT_ROLES = _parse_id_set(DISCORD_EXPORT_ROLE_IDS)

def _now_kst() -> datetime:
    return datetime.now(tz=KST)

def _today_kst() -> date:
    return _now_kst().date()


def _normalize_place_nickname(raw: str) -> str:
    text = re.sub(r"[\x00-\x1f\x7f]", "", raw or "")
    text = re.sub(r"\s+", " ", text).strip()
    return text[:50]


def _escape_discord_text(raw: str) -> str:
    text = re.sub(r"[\x00-\x1f\x7f]", "", raw or "")
    return text.replace("@", "@\u200b").strip()


def _normalize_loc(raw: str) -> str:
    text = (raw or "").strip()
    text = re.sub(r"[^a-zA-Z0-9가-힣_-]", "", text)
    return text[:64]


def _store_oauth_state(session: dict, loc: str) -> str:
    state = secrets.token_urlsafe(32)
    session["oauth_state"] = {
        "value": state,
        "loc": loc,
        "ts": int(_now_kst().timestamp()),
    }
    return state


def _pop_valid_oauth_state(session: dict, state: str) -> Optional[str]:
    data = session.pop("oauth_state", None)
    if not data or not isinstance(data, dict):
        return None
    expected = str(data.get("value") or "")
    loc = _normalize_loc(str(data.get("loc") or ""))
    ts = int(data.get("ts") or 0)
    if not expected or not state or not secrets.compare_digest(expected, state):
        return None
    if ts <= 0 or int(_now_kst().timestamp()) - ts > OAUTH_STATE_TTL_SECONDS:
        return None
    return loc

# ----------------------------
# Middleware: session TTL
# ----------------------------
@app.middleware("http")
async def _session_ttl_middleware(request: Request, call_next):
    try:
        if request.session.get("user") and request.session.get("login_ts"):
            ts = int(request.session.get("login_ts") or 0)
            if ts > 0 and int(_now_kst().timestamp()) - ts > WEB_SESSION_TTL_SECONDS:
                request.session.clear()
    except Exception:
        pass
    return await call_next(request)

# ----------------------------
# Database schema
# ----------------------------
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS places (
  id BIGSERIAL PRIMARY KEY,
  slug TEXT UNIQUE NOT NULL,
  nickname TEXT NOT NULL,
  created_by BIGINT NOT NULL,
  created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  admin_channel_id BIGINT,
  visit_role_id BIGINT
);

ALTER TABLE places ADD COLUMN IF NOT EXISTS admin_channel_id BIGINT;
ALTER TABLE places ADD COLUMN IF NOT EXISTS visit_role_id BIGINT;

CREATE TABLE IF NOT EXISTS visits (
  id BIGSERIAL PRIMARY KEY,
  place_id BIGINT NOT NULL REFERENCES places(id) ON DELETE CASCADE,
  user_id BIGINT NOT NULL,
  visit_date DATE NOT NULL,
  visit_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
  server_display_name TEXT,
  username TEXT,
  UNIQUE(place_id, user_id, visit_date)
);

CREATE INDEX IF NOT EXISTS idx_visits_place_date ON visits(place_id, visit_date);
CREATE INDEX IF NOT EXISTS idx_visits_user_date ON visits(user_id, visit_date);
"""

async def db() -> asyncpg.Pool:
    global _db_pool
    if _db_pool is None:
        if not DATABASE_URL:
            raise RuntimeError("DATABASE_URL is not set")
        _db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)
        async with _db_pool.acquire() as conn:
            await conn.execute(SCHEMA_SQL)
    return _db_pool

# ----------------------------
# Discord API helper
# ----------------------------
async def discord_api(
    method: str,
    path: str,
    *,
    bot: bool = True,
    token: Optional[str] = None,
    json_body: Optional[dict] = None,
) -> httpx.Response:
    if bot:
        tok = DISCORD_BOT_TOKEN
    else:
        tok = token or ""
    headers = {
        "Authorization": f"Bot {tok}" if bot else f"Bearer {tok}",
        "User-Agent": "qr-checkin-bot (https://example.invalid, 1.0)",
    }
    if json_body is not None:
        headers["Content-Type"] = "application/json"

    url = f"https://discord.com/api/v10{path}"
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.request(method, url, headers=headers, json=json_body)
    return r

async def get_guild_member(user_id: int) -> Optional[dict]:
    if not DISCORD_GUILD_ID:
        return None
    r = await discord_api("GET", f"/guilds/{DISCORD_GUILD_ID}/members/{user_id}", bot=True)
    if r.status_code == 200:
        return r.json()
    return None

def _member_display_name(member: dict) -> str:
    # Prefer server nickname, fallback to global_name/username
    nick = member.get("nick")
    if nick:
        return nick
    u = member.get("user") or {}
    return u.get("global_name") or u.get("username") or str(u.get("id") or "")

def _member_username(member: dict) -> str:
    u = member.get("user") or {}
    # Prefer "username#discriminator" if discriminator exists (legacy), else username.
    username = u.get("username") or ""
    disc = u.get("discriminator")
    if disc and disc != "0":
        return f"{username}#{disc}"
    return username or str(u.get("id") or "")

async def get_guild_roles_map() -> Dict[int, str]:
    # Cache for 60 seconds
    now_ts = _now_kst().timestamp()
    if _roles_cache["map"] and now_ts - float(_roles_cache["ts"] or 0) < 60:
        return _roles_cache["map"]

    if not DISCORD_GUILD_ID:
        return {}
    r = await discord_api("GET", f"/guilds/{DISCORD_GUILD_ID}/roles", bot=True)
    if r.status_code != 200:
        return _roles_cache.get("map") or {}
    roles = r.json()
    m = {int(x["id"]): x.get("name", "") for x in roles if x.get("id")}
    _roles_cache["ts"] = now_ts
    _roles_cache["map"] = m
    return m

async def send_dm(user_id: int, content: str):
    # Create DM channel
    r = await discord_api("POST", "/users/@me/channels", bot=True, json_body={"recipient_id": str(user_id)})
    if r.status_code != 200:
        return
    ch = r.json()
    ch_id = ch.get("id")
    if not ch_id:
        return
    await discord_api("POST", f"/channels/{ch_id}/messages", bot=True, json_body={"content": content})

async def send_channel_message(
    channel_id: int,
    content: str,
    components: Optional[list] = None,
    allowed_mentions: Optional[dict] = None,
):
    body: Dict[str, Any] = {"content": content}
    if components is not None:
        body["components"] = components
    if allowed_mentions is not None:
        body["allowed_mentions"] = allowed_mentions
    await discord_api("POST", f"/channels/{channel_id}/messages", bot=True, json_body=body)

async def send_channel_message_return(
    channel_id: int,
    content: str,
    components: Optional[list] = None,
    allowed_mentions: Optional[dict] = None,
) -> Optional[dict]:
    """Send a message and return the created message payload (or None on failure)."""
    body: Dict[str, Any] = {"content": content}
    if components is not None:
        body["components"] = components
    if allowed_mentions is not None:
        body["allowed_mentions"] = allowed_mentions
    r = await discord_api("POST", f"/channels/{channel_id}/messages", bot=True, json_body=body)
    if r.status_code in (200, 201):
        try:
            return r.json()
        except Exception:
            return None
    return None

async def edit_channel_message(
    channel_id: int,
    message_id: int,
    *,
    content: Optional[str] = None,
    components: Optional[list] = None,
    allowed_mentions: Optional[dict] = None,
):
    """Edit an existing message."""
    body: Dict[str, Any] = {}
    if content is not None:
        body["content"] = content
    if components is not None:
        body["components"] = components
    if allowed_mentions is not None:
        body["allowed_mentions"] = allowed_mentions
    await discord_api("PATCH", f"/channels/{channel_id}/messages/{message_id}", bot=True, json_body=body)


# ----------------------------
# Permissions helpers
# ----------------------------
def _member_role_ids_from_interaction(interaction: dict) -> Set[int]:
    roles = (interaction.get("member") or {}).get("roles") or []
    out: Set[int] = set()
    for rid in roles:
        try:
            out.add(int(rid))
        except Exception:
            pass
    return out

def _is_admin_interaction(interaction: dict) -> bool:
    # Administrator permission bit (0x8) in the "permissions" string for members.
    perms_str = (interaction.get("member") or {}).get("permissions") or "0"
    try:
        perms = int(perms_str)
        if perms & 0x8:
            return True
    except Exception:
        pass
    if not ADMIN_ROLES:
        return False
    return bool(_member_role_ids_from_interaction(interaction) & ADMIN_ROLES)

def _can_create_qr_interaction(interaction: dict) -> bool:
    # QR 생성 전용 권한: QR_CREATE_ROLES 또는 관리자
    if _is_admin_interaction(interaction):
        return True
    if not QR_CREATE_ROLES:
        return False
    return bool(_member_role_ids_from_interaction(interaction) & QR_CREATE_ROLES)

def _can_export_interaction(interaction: dict) -> bool:
    # 방문 로그 엑셀 내보내기 권한: EXPORT_ROLES 또는 관리자
    if _is_admin_interaction(interaction):
        return True
    if not EXPORT_ROLES:
        return False
    return bool(_member_role_ids_from_interaction(interaction) & EXPORT_ROLES)

# ----------------------------
# Utility
# ----------------------------
def safe_slug(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-zA-Z0-9가-힣_-]", "", s)
    return s[:64] or "place"

def base_url(request: Request) -> str:
    return PUBLIC_BASE_URL

# ----------------------------
# OAuth helpers
# ----------------------------
def discord_oauth_authorize_url(state: str) -> str:
    # identify is necessary to know who is checking in.
    scope = "identify"
    params = {
        "client_id": DISCORD_CLIENT_ID,
        "redirect_uri": OAUTH_REDIRECT_URI,
        "response_type": "code",
        "scope": scope,
        "state": state,
    }
    from urllib.parse import urlencode
    return "https://discord.com/oauth2/authorize?" + urlencode(params)

async def exchange_code(code: str) -> dict:
    data = {
        "client_id": DISCORD_CLIENT_ID,
        "client_secret": DISCORD_CLIENT_SECRET,
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": OAUTH_REDIRECT_URI,
    }
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.post("https://discord.com/api/v10/oauth2/token", data=data, headers=headers)
        r.raise_for_status()
        return r.json()

async def fetch_user(token: str) -> dict:
    r = await discord_api("GET", "/users/@me", bot=False, token=token)
    r.raise_for_status()
    return r.json()

# ----------------------------
# Startup: register slash commands
# ----------------------------
@app.on_event("startup")
async def startup():
    await db()
    await ensure_commands()

async def ensure_commands():
    """
    Register/update application commands in the target guild for faster propagation.
    """
    if not (DISCORD_CLIENT_ID and DISCORD_BOT_TOKEN and DISCORD_GUILD_ID):
        return

    commands = [
        {
            "name": "qr생성",
            "description": "새 장소 QR을 생성합니다 (QR 이미지는 바로 표시되지 않음)",
            "options": [
                {"name": "닉네임", "description": "장소 표시 이름", "type": 3, "required": True},
                {"name": "슬러그", "description": "URL에 들어갈 키 (영문/숫자/하이픈 권장)", "type": 3, "required": False},
                {"name": "운영채널", "description": "운영진 알림을 보낼 채널", "type": 7, "required": False},
            ],
        },
        {
            "name": "qr닉네임수정",
            "description": "생성된 QR(장소)의 닉네임을 수정합니다",
        },
        {"name": "qr조회", "description": "생성된 QR을 조회합니다 (명령어 사용 시에만 DM/에페메랄로 표시)"},
        {"name": "qr삭제", "description": "생성된 QR을 삭제합니다"},
        {"name": "유저방문기록", "description": "특정 유저의 전체 방문 기록을 조회합니다 (관리자 전용)", "options": [
            {"name": "유저id", "description": "Discord User ID", "type": 3, "required": True},
        ]},
        {"name": "장소방문기록", "description": "특정 장소의 유저별 방문 기록을 조회합니다 (페이지네이션)"},
        {"name": "체크인초기화", "description": "특정 유저의 특정 장소 체크인 타이머(오늘 방문 기록) 초기화"},
        {"name": "방문기록삭제", "description": "특정 유저의 특정 장소 방문 기록 전체 삭제"},
        {"name": "장소일자별방문내역", "description": "특정 장소를 기간으로 조회해 유저별 방문횟수를 원형 그래프로 표시합니다", "options": [
            {"name": "시작일", "description": "YYYYMMDD", "type": 3, "required": True},
            {"name": "종료일", "description": "YYYYMMDD", "type": 3, "required": True},
        ]},
        {"name": "방문로그엑셀", "description": "전체 방문 로그를 엑셀(.xlsx)로 내보냅니다 (권한 필요)"},
    ]

    # Upsert: PUT all commands
    await discord_api(
        "PUT",
        f"/applications/{DISCORD_CLIENT_ID}/guilds/{DISCORD_GUILD_ID}/commands",
        bot=True,
        json_body=commands,
    )

# ----------------------------
# Web routes
# ----------------------------
@app.get("/", response_class=HTMLResponse)
async def index(request: Request, loc: str = ""):
    # If loc is given, store it (so /oauth/callback can return properly)
    loc = _normalize_loc(loc or request.session.get("loc") or "")
    if loc:
        request.session["loc"] = loc

    place_name = "장소"
    status_text = "로그인 필요"
    logout_style = "display:none"

    if loc:
        pool = await db()
        async with pool.acquire() as conn:
            row = await conn.fetchrow("SELECT nickname FROM places WHERE slug=$1", loc)
            if row:
                place_name = row["nickname"]
            else:
                place_name = "등록되지 않은 장소"

    is_logged_in = bool(request.session.get("user"))
    if is_logged_in:
        status_text = "로그인됨"
        logout_style = ""

    oauth_url = "/oauth/start"
    if loc:
        oauth_url += f"?loc={loc}"

    # Mobile "APP login" deep-link attempts work best with an absolute Discord authorize URL.
    # (Using a relative URL like "/oauth/start" breaks URL parsing in many mobile browsers.)
    oauth_state = _store_oauth_state(request.session, loc) if loc else ""
    discord_authorize_url = discord_oauth_authorize_url(oauth_state) if oauth_state else ""

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "loc": loc,
            "place_name": place_name,
            "status_text": status_text,
            "logout_style": logout_style,
            "is_logged_in": is_logged_in,
            "oauth_url": oauth_url,
            "discord_authorize_url": discord_authorize_url,
        },
    )

@app.get("/oauth/start")
async def oauth_start(request: Request, loc: str = ""):
    loc = _normalize_loc(loc or request.session.get("loc") or "")
    if not loc:
        return RedirectResponse("/", status_code=302)

    request.session["loc"] = loc
    state = _store_oauth_state(request.session, loc)
    return RedirectResponse(discord_oauth_authorize_url(state), status_code=302)

@app.get("/oauth/callback")
async def oauth_callback(request: Request, code: str = "", state: str = ""):
    valid_loc = _pop_valid_oauth_state(request.session, state)
    fallback_loc = _normalize_loc(request.session.get("loc") or "")

    if not code:
        return RedirectResponse(f"/?loc={fallback_loc}", status_code=302)
    if not valid_loc:
        request.session.clear()
        return RedirectResponse("/", status_code=302)

    loc = valid_loc

    try:
        tok = await exchange_code(code)
        access_token = tok.get("access_token")
        if not access_token:
            raise RuntimeError("no access_token")
        user = await fetch_user(access_token)

        request.session["user"] = {
            "id": int(user.get("id")),
            "username": user.get("username") or "",
        }
        request.session["login_ts"] = int(_now_kst().timestamp())
    except Exception:
        request.session.clear()
        return RedirectResponse(f"/?loc={loc}", status_code=302)

    return RedirectResponse(f"/?loc={loc}", status_code=302)

@app.get("/logout")
async def logout(request: Request):
    loc = (request.session.get("loc") or "").strip()
    request.session.clear()
    return RedirectResponse(f"/?loc={loc}", status_code=302)

# ----------------------------
# Check-in API
# ----------------------------

@app.post("/api/checkin")
async def api_checkin(request: Request):
    user = request.session.get("user")
    if not user:
        return JSONResponse({"detail": "로그인이 필요합니다."}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        body = {}
    loc = _normalize_loc((body.get("loc") if isinstance(body, dict) else "") or request.session.get("loc") or "")
    if not loc:
        return JSONResponse({"detail": "loc가 없습니다."}, status_code=400)

    user_id = int(user.get("id"))
    now = _now_kst()
    today = now.date()

    pool = await db()
    async with pool.acquire() as conn:
        place = await conn.fetchrow("SELECT id, nickname, admin_channel_id FROM places WHERE slug=$1", loc)
        if not place:
            return JSONResponse({"detail": "등록되지 않은 장소입니다."}, status_code=404)

        place_id = int(place["id"])
        place_nickname = _normalize_place_nickname(place["nickname"] or "")
        admin_channel_id = int(place["admin_channel_id"] or 0)


        # Try insert today's visit
        try:
            rec = await conn.fetchrow(
                "INSERT INTO visits(place_id, user_id, visit_date, server_display_name, username) VALUES($1,$2,$3,$4,$5) RETURNING id, visit_at",
                place_id,
                user_id,
                today,
                None,
                None,
            )
            inserted = True
            visit_id = int(rec["id"])
            visit_at = rec["visit_at"]
        except asyncpg.UniqueViolationError:
            inserted = False
            visit_id = 0
            visit_at = now

        if not inserted:
            # Already checked-in today: do NOT DM or admin-log
            return JSONResponse({"ok": True, "message": "이미 오늘 체크인했습니다. (하루 1회)"})

        # Fetch server display name + roles using bot (more reliable than OAuth for your purpose)
        server_display = ""
        username = str(user_id)
        member = await get_guild_member(user_id)
        if member:
            server_display = _member_display_name(member)
            username = _member_username(member)

        # Save names
        await conn.execute(
            "UPDATE visits SET server_display_name=$1, username=$2 WHERE id=$3",
            server_display or None,
            username or None,
            visit_id,
        )

        # Total visits for this user at this place (after insert)
        total_visits = await conn.fetchval("SELECT COUNT(*) FROM visits WHERE place_id=$1 AND user_id=$2", place_id, user_id)
        total_visits = int(total_visits or 0)

    # DM
    dm_name = _escape_discord_text(server_display or username)
    place_nickname_safe = _escape_discord_text(place_nickname)
    await send_dm(
        user_id,
        f"{dm_name}님 {place_nickname_safe} 방문을 환영합니다! (누적 {total_visits}번째 방문이시네요)",
    )

    # Admin log (per-place channel preferred, fallback to global)
    fallback_channel_id = int(DISCORD_ADMIN_CHANNEL_ID or 0)
    target_channel_id = admin_channel_id or fallback_channel_id

    if target_channel_id:
        # --- No-ping role mention workaround ---
        # 1) Send a "plain roles" message first (no role mentions).
        content_plain, components_plain = await _checkin_message_data(visit_id, page=1, roles_style="plain")
        msg = await send_channel_message_return(
            target_channel_id,
            content_plain,
            components=components_plain,
            allowed_mentions={"parse": ["users"]},
        )

        # 2) Immediately edit the message to include real role mentions.
        # Editing is widely observed to avoid triggering mention notifications in most clients.
        content_mention, components = await _checkin_message_data(visit_id, page=1, roles_style="mention")
        if msg and msg.get("id"):
            try:
                await edit_channel_message(
                    target_channel_id,
                    int(msg["id"]),
                    content=content_mention,
                    components=components,
                    allowed_mentions=_allowed_mentions_for_checkin(content_mention),
                )
            except Exception:
                # Fallback: if edit fails, send a second message with mentions.
                await send_channel_message(
                    target_channel_id,
                    content_mention,
                    components=components,
                    allowed_mentions=_allowed_mentions_for_checkin(content_mention),
                )
        else:
            # Fallback if we couldn't get a message id
            await send_channel_message(
                target_channel_id,
                content_mention,
                components=components,
                allowed_mentions=_allowed_mentions_for_checkin(content_mention),
            )

    return JSONResponse({"ok": True, "message": f"{place_nickname} 체크인 완료! (하루 1회)"})

# ----------------------------
# QR Code image endpoint (used by /qr조회)
# ----------------------------
import qrcode
from io import BytesIO

@app.get("/qr/{slug}.png")
async def qr_png(request: Request, slug: str):
    loc = _normalize_loc(slug)
    url = f"{base_url(request)}/?loc={loc}"
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png")

# ----------------------------
# Discord Interactions endpoint (slash commands + components)
# ----------------------------
def verify_discord_signature(headers: dict, body: bytes) -> bool:
    try:
        sig = headers.get("x-signature-ed25519")
        ts = headers.get("x-signature-timestamp")
        if not sig or not ts or not DISCORD_PUBLIC_KEY:
            return False
        try:
            ts_int = int(ts)
        except Exception:
            return False
        now_utc = int(datetime.now(timezone.utc).timestamp())
        if abs(now_utc - ts_int) > DISCORD_SIGNATURE_MAX_AGE_SECONDS:
            return False
        verify_key = VerifyKey(bytes.fromhex(DISCORD_PUBLIC_KEY))
        verify_key.verify(ts.encode() + body, bytes.fromhex(sig))
        return True
    except BadSignatureError:
        return False
    except Exception:
        return False

def interaction_response(content: str, ephemeral: bool = True, components: Optional[list] = None, embeds: Optional[list] = None) -> dict:
    data: Dict[str, Any] = {"content": content}
    if ephemeral:
        data["flags"] = 1 << 6
    if components is not None:
        data["components"] = components
    if embeds is not None:
        data["embeds"] = embeds
    return {"type": 4, "data": data}

def update_message_response(
    content: Optional[str] = None,
    components: Optional[list] = None,
    embeds: Optional[list] = None,
    allowed_mentions: Optional[dict] = None,
) -> dict:
    data: Dict[str, Any] = {}
    if content is not None:
        data["content"] = content
    if components is not None:
        data["components"] = components
    if embeds is not None:
        data["embeds"] = embeds
    if allowed_mentions is not None:
        data["allowed_mentions"] = allowed_mentions
    return {"type": 7, "data": data}

def modal_response(custom_id: str, title: str, components: list) -> dict:
    return {"type": 9, "data": {"custom_id": custom_id, "title": title, "components": components}}

def _place_select_component(custom_id: str, options: List[dict], placeholder: str = "장소 선택") -> list:
    return [
        {
            "type": 1,
            "components": [
                {
                    "type": 3,
                    "custom_id": custom_id,
                    "placeholder": placeholder,
                    "options": options[:25],
                    "min_values": 1,
                    "max_values": 1,
                }
            ],
        }
    ]

async def _places_options_for_user(user_id: int, is_admin: bool) -> List[dict]:
    pool = await db()
    async with pool.acquire() as conn:
        if is_admin:
            rows = await conn.fetch("SELECT id, nickname FROM places ORDER BY id DESC")
        else:
            rows = await conn.fetch("SELECT id, nickname FROM places WHERE created_by=$1 ORDER BY id DESC", user_id)
    opts = []
    for r in rows:
        opts.append({"label": r["nickname"], "value": str(r["id"])})
    if not opts:
        opts.append({"label": "(생성된 장소가 없습니다)", "value": "0", "description": "먼저 /qr생성으로 장소를 만들어요"})
    return opts

async def _get_place(place_id: int) -> Optional[dict]:
    pool = await db()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM places WHERE id=$1", place_id)
        return dict(row) if row else None

async def _can_manage_place(user_id: int, interaction: dict, place_id: int) -> bool:
    if _is_admin_interaction(interaction):
        return True
    place = await _get_place(place_id)
    if not place:
        return False
    return int(place.get("created_by") or 0) == int(user_id)

async def _get_visit_place_id(visit_id: int) -> int:
    pool = await db()
    async with pool.acquire() as conn:
        place_id = await conn.fetchval("SELECT place_id FROM visits WHERE id=$1", int(visit_id))
    return int(place_id or 0)



def _get_option_value(options: list, name: str) -> Optional[Any]:
    for opt in options or []:
        if opt.get("name") == name:
            return opt.get("value")
    return None

def _parse_yyyymmdd(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not re.fullmatch(r"\d{8}", s):
        return None
    try:
        return datetime.strptime(s, "%Y%m%d").date()
    except Exception:
        return None

def _quickchart_pie_url(labels: List[str], values: List[int]) -> str:
    # Use QuickChart (Discord will fetch it)
    config = {
        "type": "pie",
        "data": {"labels": labels, "datasets": [{"data": values}]},
        "options": {
            "plugins": {"legend": {"position": "right"}},
        },
    }
    c = json.dumps(config, ensure_ascii=False, separators=(",", ":"))
    from urllib.parse import quote
    return "https://quickchart.io/chart?c=" + quote(c)

def _nav_buttons(custom_prefix: str, page: int, total_pages: int) -> list:
    prev_disabled = page <= 1
    next_disabled = page >= total_pages
    return [
        {
            "type": 1,
            "components": [
                {"type": 2, "style": 2, "label": "◀", "custom_id": f"{custom_prefix}:{page-1}", "disabled": prev_disabled},
                {"type": 2, "style": 2, "label": f"{page}/{total_pages}", "custom_id": "noop", "disabled": True},
                {"type": 2, "style": 2, "label": "▶", "custom_id": f"{custom_prefix}:{page+1}", "disabled": next_disabled},
            ],
        }
    ]


# ----------------------------
# Allowed mentions helper
# ----------------------------
_ROLE_MENTION_RE = re.compile(r"<@&(\d+)>")

def _allowed_mentions_for_checkin(content: str) -> dict:
    """Allow user mentions, and allow ONLY the role mentions that appear in the content."""
    role_ids = [m for m in _ROLE_MENTION_RE.findall(content or "")]
    # NOTE: allowing role mentions will ping members of mentionable roles.
    # If you want zero pings, keep roles as plain text instead of mentions.
    return {"parse": ["users"], "roles": role_ids}

async def _render_daily_pie(place_id: int, start: date, end: date, page: int) -> Tuple[dict, list]:
    # Returns (embed, components)
    pool = await db()
    async with pool.acquire() as conn:
        place = await conn.fetchrow("SELECT nickname FROM places WHERE id=$1", place_id)
        if not place:
            raise ValueError("place not found")
        place_name = _normalize_place_nickname(place["nickname"] or "")

        rows = await conn.fetch(
            """
            SELECT user_id,
                   COALESCE(MAX(server_display_name), '') AS sd,
                   COALESCE(MAX(username), '') AS un,
                   COUNT(*) AS cnt
            FROM visits
            WHERE place_id=$1 AND visit_date BETWEEN $2 AND $3
            GROUP BY user_id
            ORDER BY cnt DESC
            """,
            place_id,
            start,
            end,
        )

    items = []
    for r in rows:
        uid = int(r["user_id"])
        sd = (r["sd"] or "").strip()
        un = (r["un"] or "").strip()
        name = sd or un or str(uid)
        if sd and un and sd != un:
            label = f"{sd} ({un})"
        else:
            label = name
        # keep labels short for chart readability
        if len(label) > 24:
            label = label[:23] + "…"
        items.append((label, int(r["cnt"])))

    if not items:
        embed = {
            "title": f"{place_name} · 일자별 방문내역",
            "description": f"{start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}\n\n표시할 방문 기록이 없습니다.",
        }
        return embed, []

    page_size = 10
    total_pages = max(1, math.ceil(len(items) / page_size))
    page = max(1, min(page, total_pages))
    seg = items[(page - 1) * page_size : page * page_size]

    labels = [x[0] for x in seg]
    values = [x[1] for x in seg]
    chart_url = _quickchart_pie_url(labels, values)

    legend_lines = []
    for i, (lbl, cnt) in enumerate(seg, 1):
        legend_lines.append(f"{i}. {lbl} — {cnt}회")

    embed = {
        "title": f"{place_name} · 일자별 방문내역",
        "description": f"{start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}\n\n" + "\n".join(legend_lines),
        "image": {"url": chart_url},
    }

    prefix = f"daily_pie_page:{place_id}:{start.strftime('%Y%m%d')}:{end.strftime('%Y%m%d')}"
    components = _nav_buttons(prefix, page, total_pages) if total_pages > 1 else []
    return embed, components

# ----------------------------
# Discord follow-up helpers (for file exports)
# ----------------------------
def defer_response(ephemeral: bool = True) -> dict:
    data: Dict[str, Any] = {}
    if ephemeral:
        data["flags"] = 1 << 6
    return {"type": 5, "data": data}

async def send_interaction_followup_file(
    application_id: str,
    interaction_token: str,
    *,
    filename: str,
    file_bytes: bytes,
    content: str = "",
    ephemeral: bool = True,
) -> bool:
    """Send a follow-up message (supports file upload) using the interaction webhook."""
    url = f"https://discord.com/api/v10/webhooks/{application_id}/{interaction_token}"
    payload: Dict[str, Any] = {"content": content or ""}
    if ephemeral:
        payload["flags"] = 1 << 6

    files = {
        "files[0]": (
            filename,
            file_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    data = {"payload_json": json.dumps(payload, ensure_ascii=False)}

    async with httpx.AsyncClient(timeout=60.0) as client:
        r = await client.post(url, data=data, files=files)
    return r.status_code in (200, 201, 204)

def _sanitize_sheet_title(title: str) -> str:
    t = (title or "").strip() or "Sheet"
    t = re.sub(r"[:\\/?*\[\]]", "_", t)
    return t[:31] or "Sheet"

def _autosize_worksheet(ws):
    # Lightweight autosize (caps at 50)
    for col in ws.columns:
        try:
            col_letter = get_column_letter(col[0].column)
        except Exception:
            continue
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def build_visits_workbook(rows: List[dict]) -> bytes:
    """Build an .xlsx bytes payload from visit rows (joined with places)."""
    header_font = Font(bold=True)
    center = Alignment(vertical="center")

    # Aggregate
    by_place: Dict[str, List[dict]] = {}
    for r in rows:
        place = (r.get("place") or "").strip() or "(미지정)"
        by_place.setdefault(place, []).append(r)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"
    ws_summary.append(["장소", "총 방문수", "고유 방문자", "마지막 방문(KST)"])
    for c in ws_summary[1]:
        c.font = header_font
        c.alignment = center
    ws_summary.freeze_panes = "A2"

    def _fmt_kst(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")

    # Summary rows
    for place, items in sorted(by_place.items(), key=lambda x: x[0].lower()):
        total_visits = len(items)
        uniq = len({int(i.get("user_id") or 0) for i in items})
        last_dt = None
        for i in items:
            dt = i.get("visit_at")
            if isinstance(dt, datetime):
                if last_dt is None or dt > last_dt:
                    last_dt = dt
        ws_summary.append([place, total_visits, uniq, _fmt_kst(last_dt)])

    _autosize_worksheet(ws_summary)

    # All records sheet
    ws_all = wb.create_sheet("전체기록")
    ws_all.append(["장소", "닉네임", "User ID", "방문 시각(KST)"])
    for c in ws_all[1]:
        c.font = header_font
        c.alignment = center
    ws_all.freeze_panes = "A2"

    # Sort by time desc
    def _row_time(r: dict):
        dt = r.get("visit_at")
        if isinstance(dt, datetime):
            return dt
        return datetime.min.replace(tzinfo=timezone.utc)

    rows_sorted = sorted(rows, key=_row_time, reverse=True)
    for r in rows_sorted:
        place = (r.get("place") or "").strip() or "(미지정)"
        uid = int(r.get("user_id") or 0)
        nick = (r.get("display") or "").strip() or str(uid)
        dt = r.get("visit_at")
        ws_all.append([place, nick, uid, _fmt_kst(dt if isinstance(dt, datetime) else None)])

    _autosize_worksheet(ws_all)

    # Place sheets: per-user aggregates
    for place, items in sorted(by_place.items(), key=lambda x: x[0].lower()):
        ws = wb.create_sheet(_sanitize_sheet_title(place))
        ws.append(["닉네임", "User ID", "누적 방문횟수", "마지막 방문(KST)"])
        for c in ws[1]:
            c.font = header_font
            c.alignment = center
        ws.freeze_panes = "A2"

        agg: Dict[int, Dict[str, Any]] = {}
        for it in items:
            uid = int(it.get("user_id") or 0)
            nick = (it.get("display") or "").strip() or str(uid)
            dt = it.get("visit_at")
            if uid not in agg:
                agg[uid] = {"nick": nick, "count": 0, "last": None}
            agg[uid]["count"] += 1
            if isinstance(dt, datetime):
                if agg[uid]["last"] is None or dt > agg[uid]["last"]:
                    agg[uid]["last"] = dt

        # Sort by count desc then nick
        for uid, a in sorted(agg.items(), key=lambda x: (-int(x[1]["count"]), str(x[1]["nick"]).lower())):
            ws.append([a["nick"], uid, int(a["count"]), _fmt_kst(a.get("last"))])

        _autosize_worksheet(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

async def _task_export_visits_xlsx(interaction: dict) -> None:
    """Background task: build visits xlsx and send it as an interaction follow-up."""
    token = str(interaction.get("token") or "")
    app_id = str(interaction.get("application_id") or DISCORD_CLIENT_ID or "")
    if not token or not app_id:
        return

    try:
        pool = await db()
        async with pool.acquire() as conn:
            raw_rows = await conn.fetch(
                """
                SELECT p.nickname AS place,
                       v.user_id,
                       COALESCE(v.server_display_name, '') AS server_display_name,
                       COALESCE(v.username, '') AS username,
                       v.visit_at
                FROM visits v
                JOIN places p ON p.id=v.place_id
                ORDER BY v.visit_at DESC
                LIMIT 50000
                """
            )
    except Exception:
        await send_interaction_followup_file(
            app_id,
            token,
            filename="visits_error.txt",
            file_bytes="DB 조회 중 오류가 발생했습니다.".encode("utf-8"),
            content="❌ 방문 로그를 불러오지 못했어요.",
            ephemeral=True,
        )
        return

    rows: List[dict] = []
    for r in raw_rows:
        uid = int(r.get("user_id") or 0)
        sd = (r.get("server_display_name") or "").strip()
        un = (r.get("username") or "").strip()
        display = sd or un or str(uid)
        rows.append(
            {
                "place": r.get("place") or "",
                "user_id": uid,
                "display": display,
                "visit_at": r.get("visit_at"),
            }
        )

    try:
        xlsx_bytes = build_visits_workbook(rows)
        ts = _now_kst().strftime("%Y%m%d_%H%M%S")
        filename = f"visits_{ts}_KST.xlsx"
        ok = await send_interaction_followup_file(
            app_id,
            token,
            filename=filename,
            file_bytes=xlsx_bytes,
            content="✅ 방문 로그 엑셀 파일입니다.",
            ephemeral=True,
        )
        if not ok:
            await send_interaction_followup_file(
                app_id,
                token,
                filename="visits_error.txt",
                file_bytes="파일 업로드에 실패했습니다.".encode("utf-8"),
                content="❌ 엑셀 파일 업로드에 실패했어요.",
                ephemeral=True,
            )
    except Exception:
        await send_interaction_followup_file(
            app_id,
            token,
            filename="visits_error.txt",
            file_bytes="엑셀 생성 중 오류가 발생했습니다.".encode("utf-8"),
            content="❌ 엑셀 생성 중 오류가 발생했어요.",
            ephemeral=True,
        )

@app.post("/discord/interactions")
async def discord_interactions(request: Request):
    body = await request.body()
    if not verify_discord_signature(request.headers, body):
        return JSONResponse({"error": "invalid signature"}, status_code=401)

    interaction = json.loads(body.decode("utf-8"))
    itype = interaction.get("type")

    # PING
    if itype == 1:
        return JSONResponse({"type": 1})

    data = interaction.get("data") or {}
    user = interaction.get("member", {}).get("user") or interaction.get("user") or {}
    user_id = int(user.get("id") or 0)
    public = base_url(request)

    # Slash commands
    if itype == 2:
        name = data.get("name")
        options = data.get("options") or []
        is_admin = _is_admin_interaction(interaction)
        # ------------- /방문로그엑셀 (role-gated)
        if name == "방문로그엑셀":
            if not _can_export_interaction(interaction):
                return JSONResponse(interaction_response("이 명령어는 지정된 역할(들) 또는 관리자만 사용할 수 있어요.", ephemeral=True))

            # Defer immediately (ephemeral), then send file via follow-up webhook
            asyncio.create_task(_task_export_visits_xlsx(interaction))
            return JSONResponse(defer_response(ephemeral=True))

        # ------------- /qr생성 (role-gated)
        if name == "qr생성":
            if not _can_create_qr_interaction(interaction):
                return JSONResponse(interaction_response("이 명령어는 지정된 역할(들) 또는 관리자만 사용할 수 있어요.", ephemeral=True))

            nickname = _normalize_place_nickname(str(_get_option_value(options, "닉네임") or ""))
            slug = str(_get_option_value(options, "슬러그") or "").strip()
            channel_id_raw = _get_option_value(options, "운영채널")

            if not nickname:
                return JSONResponse(interaction_response("닉네임이 필요합니다.", ephemeral=True))

            slug = safe_slug(slug or nickname)
            admin_channel_id = 0
            if channel_id_raw is not None:
                try:
                    admin_channel_id = int(channel_id_raw)
                except Exception:
                    admin_channel_id = 0

            pool = await db()
            async with pool.acquire() as conn:
                try:
                    await conn.execute(
                        "INSERT INTO places(slug, nickname, created_by, admin_channel_id) VALUES($1,$2,$3,$4)",
                        slug,
                        nickname,
                        user_id,
                        admin_channel_id or None,
                    )
                except asyncpg.UniqueViolationError:
                    return JSONResponse(interaction_response("이미 같은 슬러그가 존재합니다. 슬러그를 바꿔주세요.", ephemeral=True))

            msg = f"✅ 생성 완료: **{nickname}**\n- URL 키: `{slug}`\n- 체크인 링크: {public}/?loc={slug}\n- QR 이미지: {public}/qr/{slug}.png (조회는 /qr조회)\n"
            if admin_channel_id:
                msg += f"- 운영 알림 채널: <#{admin_channel_id}>"
            else:
                msg += "- 운영 알림 채널: (미설정 — 전역 DISCORD_ADMIN_CHANNEL_ID 사용)"
            return JSONResponse(interaction_response(msg, ephemeral=True))

        # Remaining commands: place owner OR admins (admin roles)
        # Helper: show place selector filtered
        if name in ("qr닉네임수정", "qr조회", "qr삭제", "장소방문기록", "체크인초기화", "방문기록삭제", "장소일자별방문내역"):
            opts = await _places_options_for_user(user_id, is_admin)
            if opts and opts[0].get("value") == "0":
                return JSONResponse(interaction_response("사용 가능한 장소가 없습니다. 먼저 /qr생성으로 장소를 만들어주세요.", ephemeral=True))

            # command-specific selector custom_id
            if name == "qr닉네임수정":
                return JSONResponse(interaction_response("수정할 장소를 선택하세요.", ephemeral=True, components=_place_select_component("place_select_rename", opts)))
            if name == "qr조회":
                return JSONResponse(interaction_response("조회할 장소를 선택하세요. (QR은 에페메랄로만 표시됩니다)", ephemeral=True, components=_place_select_component("place_select_showqr", opts)))
            if name == "qr삭제":
                return JSONResponse(interaction_response("삭제할 장소를 선택하세요.", ephemeral=True, components=_place_select_component("place_select_delete", opts)))
            if name == "장소방문기록":
                return JSONResponse(interaction_response("조회할 장소를 선택하세요.", ephemeral=True, components=_place_select_component("place_select_placevisits", opts)))
            if name == "체크인초기화":
                return JSONResponse(interaction_response("초기화할 장소를 선택하세요.", ephemeral=True, components=_place_select_component("place_select_reset_timer", opts)))
            if name == "방문기록삭제":
                return JSONResponse(interaction_response("삭제할 장소를 선택하세요.", ephemeral=True, components=_place_select_component("place_select_delete_user_visits", opts)))
            if name == "장소일자별방문내역":
                start_raw = str(_get_option_value(options, "시작일") or "")
                end_raw = str(_get_option_value(options, "종료일") or "")
                sd = _parse_yyyymmdd(start_raw)
                ed = _parse_yyyymmdd(end_raw)
                if not sd or not ed:
                    return JSONResponse(interaction_response("날짜 형식이 올바르지 않습니다. YYYYMMDD로 입력해주세요.", ephemeral=True))
                if sd > ed:
                    return JSONResponse(interaction_response("시작일은 종료일보다 늦을 수 없습니다.", ephemeral=True))
                # Place select with dates in custom id
                cid = f"daily_pie_select:{sd.strftime('%Y%m%d')}:{ed.strftime('%Y%m%d')}"
                return JSONResponse(interaction_response("장소를 선택하세요.", ephemeral=True, components=_place_select_component(cid, opts)))

        # /유저방문기록 (admin-only)
        if name == "유저방문기록":
            if not is_admin:
                return JSONResponse(interaction_response("이 명령어는 관리자 전용입니다.", ephemeral=True))
            uid_raw = str(_get_option_value(options, "유저id") or "").strip()
            if not uid_raw.isdigit():
                return JSONResponse(interaction_response("유저id는 숫자여야 합니다.", ephemeral=True))
            target_uid = int(uid_raw)

            pool = await db()
            async with pool.acquire() as conn:
                rows = await conn.fetch(
                    """
                    SELECT p.nickname AS place, v.visit_date
                    FROM visits v
                    JOIN places p ON p.id=v.place_id
                    WHERE v.user_id=$1
                    ORDER BY v.visit_date DESC
                    LIMIT 50
                    """,
                    target_uid,
                )
            if not rows:
                return JSONResponse(interaction_response("기록이 없습니다.", ephemeral=True))

            lines = [f"- {r['visit_date'].strftime('%Y-%m-%d')} · {r['place']}" for r in rows]
            return JSONResponse(interaction_response("최근 50개 방문 기록:\n" + "\n".join(lines), ephemeral=True))

        return JSONResponse(interaction_response("알 수 없는 명령어입니다.", ephemeral=True))

    # Component interactions (select menus / buttons)
    if itype == 3:
        custom_id = (data.get("custom_id") or "").strip()

        # ignore noop
        if custom_id == "noop":
            return JSONResponse(update_message_response())
        # Check-in roles pagination buttons: checkin_roles:{visit_id}:{page}
        if custom_id.startswith("checkin_roles:"):
            try:
                _, vid_s, page_s = custom_id.split(":")
                vid = int(vid_s)
                page = int(page_s)
            except Exception:
                return JSONResponse(update_message_response())
            visit_place_id = await _get_visit_place_id(vid)
            if not visit_place_id or not await _can_manage_place(user_id, interaction, visit_place_id):
                return JSONResponse(interaction_response("권한이 없습니다.", ephemeral=True))
            return JSONResponse(await _render_checkin_roles_message(vid, page))


        # Place visits page buttons: place_visits_page:{place_id}:{page}
        if custom_id.startswith("place_visits_page:"):
            try:
                _, pid_s, page_s = custom_id.split(":")
                pid = int(pid_s)
                page = int(page_s)
            except Exception:
                return JSONResponse(update_message_response())
            if not await _can_manage_place(user_id, interaction, pid):
                return JSONResponse(interaction_response("권한이 없습니다.", ephemeral=True))
            return JSONResponse(await _render_place_visits_message(pid, page=page))

        # Place select handlers
        values = (data.get("values") or [])
        selected_place_id = int(values[0]) if values and str(values[0]).isdigit() else 0

        # Helper permission check
        async def _deny_if_not_manage() -> Optional[JSONResponse]:
            if not selected_place_id:
                return JSONResponse(interaction_response("장소 선택이 올바르지 않습니다.", ephemeral=True))
            if not await _can_manage_place(user_id, interaction, selected_place_id):
                return JSONResponse(interaction_response("이 장소에 대한 권한이 없습니다. (생성자 또는 관리자만 가능)", ephemeral=True))
            return None

        # Rename flow
        if custom_id == "place_select_rename":
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            return JSONResponse(
                modal_response(
                    f"modal_rename:{selected_place_id}",
                    "닉네임 수정",
                    [
                        {
                            "type": 1,
                            "components": [
                                {"type": 4, "custom_id": "new_nickname", "style": 1, "label": "새 닉네임", "min_length": 1, "max_length": 50, "required": True}
                            ],
                        }
                    ],
                )
            )

        # Show QR
        if custom_id == "place_select_showqr":
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            place = await _get_place(selected_place_id)
            if not place:
                return JSONResponse(interaction_response("장소를 찾을 수 없습니다.", ephemeral=True))

            # show QR image URL (ephemeral)
            qr_url = f"{public}/qr/{place['slug']}.png"
            page_url = f"{public}/?loc={place['slug']}"
            content = f"**{place['nickname']}**\n- QR 이미지: {qr_url}\n- 체크인 URL: {page_url}"
            return JSONResponse(interaction_response(content, ephemeral=True))

        # Delete place
        if custom_id == "place_select_delete":
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            pool = await db()
            async with pool.acquire() as conn:
                await conn.execute("DELETE FROM places WHERE id=$1", selected_place_id)
            return JSONResponse(interaction_response("✅ 삭제 완료", ephemeral=True))

        # Place visits (paged 10)
        if custom_id == "place_select_placevisits":
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            # Page 1 render
            return JSONResponse(await _render_place_visits_message(selected_place_id, page=1))

        # Reset timer (remove today's visit for a user)
        if custom_id == "place_select_reset_timer":
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            # Ask for user id via modal
            return JSONResponse(
                modal_response(
                    f"modal_reset_timer:{selected_place_id}",
                    "체크인 초기화",
                    [
                        {"type": 1, "components": [{"type": 4, "custom_id": "target_user_id", "style": 1, "label": "대상 유저ID", "min_length": 1, "max_length": 30, "required": True}]}
                    ],
                )
            )

        # Delete all visits for a user at a place
        if custom_id == "place_select_delete_user_visits":
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            return JSONResponse(
                modal_response(
                    f"modal_delete_user_visits:{selected_place_id}",
                    "방문기록 삭제",
                    [
                        {"type": 1, "components": [{"type": 4, "custom_id": "target_user_id", "style": 1, "label": "대상 유저ID", "min_length": 1, "max_length": 30, "required": True}]}
                    ],
                )
            )

        # Daily pie select: daily_pie_select:YYYYMMDD:YYYYMMDD
        if custom_id.startswith("daily_pie_select:"):
            denied = await _deny_if_not_manage()
            if denied:
                return denied
            try:
                _, s1, s2 = custom_id.split(":")
                sd = _parse_yyyymmdd(s1)
                ed = _parse_yyyymmdd(s2)
                if not sd or not ed:
                    raise ValueError
            except Exception:
                return JSONResponse(interaction_response("날짜 정보가 올바르지 않습니다.", ephemeral=True))

            embed, components = await _render_daily_pie(selected_place_id, sd, ed, page=1)
            return JSONResponse(update_message_response(content="", embeds=[embed], components=components))

        # Daily pie page buttons: daily_pie_page:{place_id}:{start}:{end}:{page}
        if custom_id.startswith("daily_pie_page:"):
            try:
                _, pid_s, s1, s2, page_s = custom_id.split(":")
                pid = int(pid_s)
                sd = _parse_yyyymmdd(s1)
                ed = _parse_yyyymmdd(s2)
                page = int(page_s)
                if not sd or not ed:
                    raise ValueError
            except Exception:
                return JSONResponse(update_message_response())

            # permission: owner/admin only
            if not await _can_manage_place(user_id, interaction, pid):
                return JSONResponse(interaction_response("권한이 없습니다.", ephemeral=True))

            embed, components = await _render_daily_pie(pid, sd, ed, page=page)
            return JSONResponse(update_message_response(embeds=[embed], components=components))

        return JSONResponse(update_message_response())

    # Modal submission
    if itype == 5:
        custom_id = (data.get("custom_id") or "").strip()

        # Helper extract modal fields
        def get_modal_value(field_id: str) -> str:
            comps = data.get("components") or []
            for row in comps:
                for c in (row.get("components") or []):
                    if c.get("custom_id") == field_id:
                        return c.get("value") or ""
            return ""

        # Rename modal
        if custom_id.startswith("modal_rename:"):
            try:
                _, pid_s = custom_id.split(":")
                pid = int(pid_s)
            except Exception:
                return JSONResponse(interaction_response("오류", ephemeral=True))

            if not await _can_manage_place(user_id, interaction, pid):
                return JSONResponse(interaction_response("권한이 없습니다.", ephemeral=True))

            new_nickname = _normalize_place_nickname(get_modal_value("new_nickname"))
            if not new_nickname:
                return JSONResponse(interaction_response("닉네임이 비어있습니다.", ephemeral=True))

            pool = await db()
            async with pool.acquire() as conn:
                await conn.execute("UPDATE places SET nickname=$1 WHERE id=$2", new_nickname, pid)

            return JSONResponse(interaction_response(f"✅ 닉네임 수정 완료: {_escape_discord_text(new_nickname)}", ephemeral=True))

        # Reset timer modal
        if custom_id.startswith("modal_reset_timer:"):
            try:
                _, pid_s = custom_id.split(":")
                pid = int(pid_s)
            except Exception:
                return JSONResponse(interaction_response("오류", ephemeral=True))

            if not await _can_manage_place(user_id, interaction, pid):
                return JSONResponse(interaction_response("권한이 없습니다.", ephemeral=True))

            target = get_modal_value("target_user_id").strip()
            if not target.isdigit():
                return JSONResponse(interaction_response("유저ID는 숫자여야 합니다.", ephemeral=True))
            target_uid = int(target)

            pool = await db()
            async with pool.acquire() as conn:
                deleted = await conn.execute(
                    "DELETE FROM visits WHERE place_id=$1 AND user_id=$2 AND visit_date=$3",
                    pid,
                    target_uid,
                    _today_kst(),
                )
            return JSONResponse(interaction_response("✅ 오늘 체크인 기록을 초기화했습니다.", ephemeral=True))

        # Delete visits modal
        if custom_id.startswith("modal_delete_user_visits:"):
            try:
                _, pid_s = custom_id.split(":")
                pid = int(pid_s)
            except Exception:
                return JSONResponse(interaction_response("오류", ephemeral=True))

            if not await _can_manage_place(user_id, interaction, pid):
                return JSONResponse(interaction_response("권한이 없습니다.", ephemeral=True))

            target = get_modal_value("target_user_id").strip()
            if not target.isdigit():
                return JSONResponse(interaction_response("유저ID는 숫자여야 합니다.", ephemeral=True))
            target_uid = int(target)

            pool = await db()
            async with pool.acquire() as conn:
                await conn.execute("DELETE FROM visits WHERE place_id=$1 AND user_id=$2", pid, target_uid)

            return JSONResponse(interaction_response("✅ 방문 기록을 삭제했습니다.", ephemeral=True))

        return JSONResponse(interaction_response("처리할 수 없는 모달입니다.", ephemeral=True))

    return JSONResponse({"error": "unhandled"}, status_code=400)

# ----------------------------
# Check-in notification roles pagination (3 roles per page)
# ----------------------------
ROLE_PAGE_SIZE = 3

def _paginate(items: List[str], page: int, page_size: int) -> Tuple[List[str], int, int]:
    total_pages = max(1, math.ceil(len(items) / page_size)) if page_size > 0 else 1
    page = max(1, min(int(page or 1), total_pages))
    start = (page - 1) * page_size
    end = start + page_size
    return items[start:end], page, total_pages

def _build_checkin_content(
    *,
    display_name: str,
    place_name: str,
    user_id: int,
    visit_time_kst: datetime,
    total_visits: int,
    roles_text: str,
    page: int,
    total_pages: int,
) -> str:
    time_str = visit_time_kst.astimezone(KST).strftime("%H:%M")
    visitor_line = f"<@{user_id}>"
    display_name = _escape_discord_text(display_name)
    place_name = _escape_discord_text(place_name)
    lines = [
        f"## [입장 알림] {display_name}님이 체크인했습니다! (누적 {total_visits}회차)",
        f"장소: {place_name}",
        f"방문자: {visitor_line}",
        f"방문 시간: {time_str} (KST)",
        f"방문 횟수: {total_visits}번째 방문",
        f"역할: {roles_text}",
        f"페이지: {page}/{total_pages}",
    ]
    return "\n".join(lines)

async def _checkin_message_data(visit_id: int, page: int, roles_style: str = "mention") -> Tuple[str, list]:
    # Resolve visit / place
    pool = await db()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT v.id, v.place_id, v.user_id, v.visit_at,
                   COALESCE(v.server_display_name, '') AS sd,
                   COALESCE(v.username, '') AS un,
                   p.nickname AS place_name
            FROM visits v
            JOIN places p ON p.id=v.place_id
            WHERE v.id=$1
            """,
            int(visit_id),
        )
        if not row:
            return ("(기록을 찾을 수 없습니다.)", [])

        place_id = int(row["place_id"])
        user_id = int(row["user_id"])
        visit_at = row["visit_at"]  # timestamptz
        place_name = _normalize_place_nickname(row["place_name"] or "")
        sd = (row["sd"] or "").strip()
        un = (row["un"] or "").strip()

        total_visits = await conn.fetchval(
            "SELECT COUNT(*) FROM visits WHERE place_id=$1 AND user_id=$2",
            place_id,
            user_id,
        )
        total_visits = int(total_visits or 0)

    # Fetch current member info (display name + roles)
    member = await get_guild_member(user_id)
    role_displays: List[str] = []
    if member:
        display_name = _member_display_name(member) or sd or un or str(user_id)
        role_ids_raw = member.get("roles") or []
        roles_map: Dict[int, str] = {}
        if roles_style != "mention":
            roles_map = await get_guild_roles_map()
        for rid in role_ids_raw:
            try:
                rid_int = int(rid)
            except Exception:
                continue
            if roles_style == "mention":
                # Real role mention (may ping if sent in a normal message create)
                role_displays.append(f"<@&{rid_int}>")
            else:
                # Plain text for the initial send (no pings). Use role name if available.
                rn = (roles_map.get(rid_int) or str(rid_int)).strip()
                if rn:
                    role_displays.append(f"@{rn}")
    else:
        display_name = sd or un or str(user_id)
        role_displays = []

    # Paginate roles 3 per page
    seg, page, total_pages = _paginate(role_displays, page, ROLE_PAGE_SIZE)
    roles_text = ", ".join(seg) if seg else "없음"

    content = _build_checkin_content(
        display_name=display_name,
        place_name=place_name,
        user_id=user_id,
        visit_time_kst=visit_at.astimezone(KST) if hasattr(visit_at, "astimezone") else _now_kst(),
        total_visits=total_visits,
        roles_text=roles_text,
        page=page,
        total_pages=total_pages,
    )

    components = _nav_buttons(f"checkin_roles:{int(visit_id)}", page, total_pages) if total_pages > 1 else []
    return content, components

async def _render_checkin_roles_message(visit_id: int, page: int) -> dict:
    content, components = await _checkin_message_data(visit_id, page)
    # Allow user mentions and role mentions that appear on the current page.
    return update_message_response(
        content=content,
        components=components,
        allowed_mentions=_allowed_mentions_for_checkin(content),
    )


# ----------------------------
# Place visits pagination (10 users per page)
# ----------------------------
async def _render_place_visits_message(place_id: int, page: int = 1) -> dict:
    pool = await db()
    async with pool.acquire() as conn:
        place = await conn.fetchrow("SELECT nickname FROM places WHERE id=$1", place_id)
        if not place:
            return interaction_response("장소를 찾을 수 없습니다.", ephemeral=True)
        place_name = _normalize_place_nickname(place["nickname"] or "")

        rows = await conn.fetch(
            """
            SELECT user_id,
                   COALESCE(MAX(server_display_name), '') AS sd,
                   COALESCE(MAX(username), '') AS un,
                   COUNT(*) AS cnt
            FROM visits
            WHERE place_id=$1
            GROUP BY user_id
            ORDER BY cnt DESC
            """,
            place_id,
        )

    items = []
    for r in rows:
        uid = int(r["user_id"])
        sd = (r["sd"] or "").strip()
        un = (r["un"] or "").strip()
        if sd and un and sd != un:
            name = f"{sd} ({un})"
        else:
            name = sd or un or str(uid)
        items.append((name, int(r["cnt"])))

    if not items:
        return interaction_response(f"**{place_name}**\n방문 기록이 없습니다.", ephemeral=True)

    page_size = 10
    total_pages = max(1, math.ceil(len(items) / page_size))
    page = max(1, min(page, total_pages))
    seg = items[(page - 1) * page_size : page * page_size]

    lines = [f"{i + (page-1)*page_size}. {name} — {cnt}회" for i, (name, cnt) in enumerate(seg, 1)]
    content = f"**{place_name}** 방문 기록 (유저별)\n\n" + "\n".join(lines) + f"\n\n페이지: {page}/{total_pages}"

    prefix = f"place_visits_page:{place_id}"
    components = _nav_buttons(prefix, page, total_pages) if total_pages > 1 else []
    # Use update response because it is called from component as well
    return update_message_response(content=content, components=components)

# Add handler for place visits page buttons in component section by reusing existing endpoint logic.
# (We intercept it via custom_id prefix in interactions above.)
